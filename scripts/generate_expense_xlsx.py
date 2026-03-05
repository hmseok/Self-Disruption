#!/usr/bin/env python3
"""
법인카드 사용내역서 엑셀 생성 스크립트
템플릿 파일을 로드하고 DB 데이터로 업데이트합니다.

사용법:
  python3 generate_expense_xlsx.py <template_path> <output_path> <month_num> <data_json_file>

  month_num: 월 번호 (예: 3)
  data_json_file: JSON 배열이 저장된 파일 경로
"""

import sys
import json
from datetime import datetime
from copy import copy
import openpyxl


def main():
    if len(sys.argv) < 5:
        print("Usage: generate_expense_xlsx.py <template> <output> <month_num> <data_json_file>", file=sys.stderr)
        sys.exit(1)

    template_path = sys.argv[1]
    output_path = sys.argv[2]
    month_num = int(sys.argv[3])
    data_json_file = sys.argv[4]

    # JSON 파일에서 데이터 읽기
    try:
        with open(data_json_file, 'r', encoding='utf-8') as f:
            items = json.load(f)
    except (json.JSONDecodeError, FileNotFoundError) as e:
        print(f"Data file error: {e}", file=sys.stderr)
        sys.exit(1)

    # 템플릿 로드
    wb = openpyxl.load_workbook(template_path)
    ws = wb['법인 지출내역서']

    # ── Row 2: 월 번호 업데이트 ──
    ws['A2'].value = f'(  {month_num}  )월 지출 합계 내역'

    # ── Row 5~34: 기존 셀 서식 보존을 위해 Row 5 기준 복사 ──
    ref_styles = {}
    for c in range(1, 9):
        cell = ws.cell(row=5, column=c)
        ref_styles[c] = {
            'font': copy(cell.font),
            'number_format': cell.number_format,
            'alignment': copy(cell.alignment),
            'border': copy(cell.border),
        }

    # 기존 데이터 삭제
    for r in range(5, 35):
        for c in range(1, 9):
            ws.cell(row=r, column=c).value = None

    # ── DB 데이터 채우기 ──
    max_rows = min(len(items), 30)  # 최대 30행

    for i in range(max_rows):
        item = items[i]
        row_num = 5 + i

        # A: 날짜
        if item.get('expense_date'):
            try:
                date_obj = datetime.strptime(item['expense_date'], '%Y-%m-%d')
                ws.cell(row=row_num, column=1).value = date_obj
            except ValueError:
                ws.cell(row=row_num, column=1).value = item['expense_date']

        # B: 카드번호
        ws.cell(row=row_num, column=2).value = item.get('card_number', '')

        # C: 구분
        ws.cell(row=row_num, column=3).value = item.get('category', '')

        # D: 사용처
        ws.cell(row=row_num, column=4).value = item.get('merchant', '')

        # E: 품명
        ws.cell(row=row_num, column=5).value = item.get('item_name', '')

        # F: 고객명/팀원
        ws.cell(row=row_num, column=6).value = item.get('customer_team', '')

        # G: 금액
        amount = item.get('amount', 0)
        if isinstance(amount, str):
            try:
                amount = int(amount.replace(',', ''))
            except ValueError:
                amount = 0
        ws.cell(row=row_num, column=7).value = amount

        # H: 영수증 첨부
        ws.cell(row=row_num, column=8).value = ''

        # 서식 적용 (템플릿 Row 5 기준)
        for c in range(1, 9):
            cell = ws.cell(row=row_num, column=c)
            style = ref_styles[c]
            cell.font = copy(style['font'])
            cell.number_format = style['number_format']
            cell.alignment = copy(style['alignment'])
            cell.border = copy(style['border'])

    # ── G2 SUM 공식 (항상 G5:G34 범위) ──
    ws['G2'].value = '=SUM(G5:G34)'

    # ── 저장 ──
    wb.save(output_path)
    print(f"OK:{output_path}")


if __name__ == '__main__':
    main()
